"""
Chuyen doi showroom/design/Table.xlsx -> showroom/design/Table.json

Chay lai script nay moi khi Table.xlsx duoc cap nhat (vd sau khi "Refresh All"):
    python Table_Converter.py

Doc 2 sheet:
  - "Khoi"    : kich thuoc tong (rong/cao/sau) cua tung khoi tu/ke.
  - "Chia O"  : danh sach tung o (vi tri, kich thuoc, va du lieu doanh nhan/san pham
                da anh xa vao o do, neu co).

Thuat toan dung luoi (dung chung cho moi khoi, khong hardcode theo tung ma):
  1. Gom cac dong Chia O theo "Thuoc ma khoi", giu nguyen thu tu STT trong file.
  2. Tach 2 nhom theo "Kieu": nhom "Tu" (cabinet, dong kin) va nhom "O trung bay"/
     "ke xien"/"khung..." (display, ho phia truoc).
  3. Dong goi trai->phai theo dung thu tu STT cho toi khi het be ngang cua khoi
     (lay tu sheet Khoi, fallback cho cac ma khong co trong Khoi nhu N/KT) roi
     xuong dong moi.
  4. Xep chong: cac dong "cabinet" o day truoc, roi den cac dong "display" o tren.
  5. Neu con du chieu cao so voi tong chieu cao khoi (lay tu sheet Khoi), them 1
     dai "hop go op tren" trang tri o dinh (kind="cap").

  Rieng nhung khoi ma moi dong deu co "Ma vi tri" == "Thuoc ma khoi" (vd KS xuat
  hien 2 dong, KT xuat hien 1 dong) duoc coi la cac don vi doc lap (khong phai
  luoi nhieu o) - moi dong la 1 block-instance rieng, khong dong goi chung.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
XLSX_PATH = SCRIPT_DIR / "Table.xlsx"
JSON_PATH = SCRIPT_DIR / "Table.json"

# Fallback tong kich thuoc (cm) cho cac ma khoi KHONG co trong sheet "Khoi"
FALLBACK_BLOCK_DIMS = {
    "N": {"width": 120, "height": 160, "depth": 50},
}


def norm(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


TRUTHY = {"true", "có", "co", "1", "x", "yes"}


def norm_bool(v):
    """Sản phẩm N có bán tại Showroom / trên HATHYO là truong Boolean."""
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    return str(v).strip().lower() in TRUTHY


def read_khoi_sheet(wb):
    ws = wb["Khối"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    blocks = {}
    for r in rows:
        if not r or not norm(r[1]):
            continue
        code = str(r[1]).strip()
        blocks[code] = {
            "code": code,
            "name": norm(r[2]),
            "mountPosition": norm(r[3]),
            "mountKind": norm(r[4]),
            "width": r[5],
            "height": r[6],
            "depth": r[7],
            "displayCellCount": norm(r[8]),
            "cabinetCellCount": norm(r[9]),
            "backing": norm(r[10]),
        }
    return blocks


def build_header_index(ws):
    header = [norm(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {name: i for i, name in enumerate(header) if name}
    return idx


def product_field_indices(idx, n):
    def find(prefix):
        key = f"{prefix} {n}"
        return idx.get(key)

    return {
        "img": find("Ảnh sản phẩm"),
        "name": find("Tên sản phẩm"),
        "desc": find("Mô tả sản phẩm"),
        "price": find("Giá đơn vị sản phẩm"),
        "qty": find("Số lượng sản phẩm"),
        "atShowroom": idx.get(f"Sản phẩm {n} có bán tại Showroom"),
        "atHathyo": idx.get(f"Sản phẩm {n} có bán trên HATHYO"),
    }


def read_chia_o_sheet(wb):
    ws = wb["Chia Ô"]
    idx = build_header_index(ws)
    prod_idx = [product_field_indices(idx, n) for n in range(1, 7)]

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    entries = []
    for r in rows:
        code = norm(r[idx["Mã vị trí"]])
        block = norm(r[idx["Thuộc mã khối"]])
        if not code or not block:
            continue

        products = []
        for pf in prod_idx:
            name = norm(r[pf["name"]]) if pf["name"] is not None else None
            img = norm(r[pf["img"]]) if pf["img"] is not None else None
            if not name and not img:
                continue
            products.append({
                "img": img,
                "name": name,
                "desc": norm(r[pf["desc"]]) if pf["desc"] is not None else None,
                "price": norm(r[pf["price"]]) if pf["price"] is not None else None,
                "qty": norm(r[pf["qty"]]) if pf["qty"] is not None else None,
                "atShowroom": norm_bool(r[pf["atShowroom"]]) if pf["atShowroom"] is not None else False,
                "atHathyo": norm_bool(r[pf["atHathyo"]]) if pf["atHathyo"] is not None else False,
            })

        entries.append({
            "code": code,
            "block": block,
            "kiểu": norm(r[idx["Kiểu"]]),
            "width": r[idx["Chiều ngang (cm)"]],
            "height": r[idx["Chiều cao (cm)"]],
            "depth": r[idx["Chiều sâu (cm)"]],
            "note": norm(r[idx["Ghi chú"]]),
            "bizCode": norm(r[idx["Mã doanh nhân"]]),
            "brand": norm(r[idx["Thương hiệu"]]),
            "company": norm(r[idx["Công ty"]]),
            "bioOwner": norm(r[idx["Giới thiệu doanh nhân"]]),
            "bioCompany": norm(r[idx["Giới thiệu công ty"]]),
            "products": products,
        })
    return entries


def is_cabinet_kind(kieu):
    return bool(kieu) and "Tủ" in kieu


STANDALONE_PREFIXES = ("standee", "mô hình", "mo hinh")


def is_standee_kind(kieu):
    """Cac vat trung bay doc lap (standee, mo hinh doi tac...) - khong dong
    goi vao luoi cua khoi, du "Thuoc ma khoi" co the trung 1 khoi tu that."""
    return bool(kieu) and kieu.strip().lower().startswith(STANDALONE_PREFIXES)


def pack_row(items, max_width):
    """Dong goi 'items' (co 'width') trai->phai thanh cac dong, moi dong <= max_width."""
    rows, cur, cur_w = [], [], 0.0
    for it in items:
        w = float(it["width"] or 0)
        if cur and cur_w + w > max_width + 0.01:
            rows.append(cur)
            cur, cur_w = [], 0.0
        cur.append(it)
        cur_w += w
    if cur:
        rows.append(cur)
    return rows


def build_blocks(khoi_blocks, chia_o_entries):
    # Standee: la vat trung bay doc lap (khong phai o trong luoi tu), du "Thuoc
    # ma khoi" co the trung voi 1 khoi tu that (vd T) - tach rieng ra truoc,
    # khong cho lot vao thuat toan dong goi luoi cua khoi do. "homeBlock" giu
    # lai de showroom.html biet dat gan tuong/khoi nao.
    standee_entries = [e for e in chia_o_entries if is_standee_kind(e["kiểu"])]
    normal_entries = [e for e in chia_o_entries if not is_standee_kind(e["kiểu"])]

    by_block = {}
    for e in normal_entries:
        by_block.setdefault(e["block"], []).append(e)

    blocks_out = {}
    cells_out = {}

    for e in standee_entries:
        cell = {
            "code": e["code"],
            "block": e["code"],
            "homeBlock": e["block"],
            "kind": "standalone",
            "row": 0,
            "x": 0,
            "y": 0,
            "width": e["width"],
            "height": e["height"],
            "depth": e["depth"],
            "note": e["note"],
            "bizCode": e["bizCode"],
            "brand": e["brand"],
            "company": e["company"],
            "bioOwner": e["bioOwner"],
            "bioCompany": e["bioCompany"],
            "products": e["products"],
        }
        cells_out[e["code"]] = cell
        blocks_out[e["code"]] = {
            "code": e["code"],
            "standalone": True,
            "homeBlock": e["block"],
            "instances": [e["code"]],
            "width": e["width"],
            "height": e["height"],
            "depth": e["depth"],
        }

    for code, items in by_block.items():
        # Truong hop dac biet: moi dong la 1 don vi doc lap (Ma vi tri == ma khoi)
        if all(it["code"] == code for it in items):
            instances = []
            for i, it in enumerate(items, start=1):
                inst_code = f"{code}-{i}" if len(items) > 1 else code
                cell = {
                    "code": inst_code,
                    "block": code,
                    "kind": "standalone",
                    "row": 0,
                    "x": 0,
                    "y": 0,
                    "width": it["width"],
                    "height": it["height"],
                    "depth": it["depth"],
                    "note": it["note"],
                    "bizCode": it["bizCode"],
                    "brand": it["brand"],
                    "company": it["company"],
                    "bioOwner": it["bioOwner"],
                    "bioCompany": it["bioCompany"],
                    "products": it["products"],
                }
                cells_out[inst_code] = cell
                instances.append(inst_code)
            blocks_out[code] = {
                "code": code,
                "standalone": True,
                "instances": instances,
                "width": items[0]["width"],
                "height": items[0]["height"],
                "depth": items[0]["depth"],
            }
            continue

        khoi = khoi_blocks.get(code)
        fallback = FALLBACK_BLOCK_DIMS.get(code, {})
        block_width = (khoi or {}).get("width") or fallback.get("width")
        block_height = (khoi or {}).get("height") or fallback.get("height")
        block_depth = (khoi or {}).get("depth") or fallback.get("depth")
        if not block_width:
            block_width = max(float(it["width"] or 0) for it in items)

        cabinet_items = [it for it in items if is_cabinet_kind(it["kiểu"])]
        display_items = [it for it in items if not is_cabinet_kind(it["kiểu"])]

        cabinet_rows = pack_row(cabinet_items, block_width)
        display_rows = pack_row(display_items, block_width)

        rows_out = []
        y = 0.0
        row_idx = 0
        for row_type, rows in (("cabinet", cabinet_rows), ("display", display_rows)):
            for row in rows:
                row_h = max(float(it["height"] or 0) for it in row)
                x = 0.0
                row_cells = []
                for it in row:
                    w = float(it["width"] or 0)
                    cell = {
                        "code": it["code"],
                        "block": code,
                        "kind": row_type,
                        "row": row_idx,
                        "x": x,
                        "y": y,
                        "width": w,
                        "height": row_h,
                        "depth": it["depth"],
                        "note": it["note"],
                        "bizCode": it["bizCode"],
                        "brand": it["brand"],
                        "company": it["company"],
                        "bioOwner": it["bioOwner"],
                        "bioCompany": it["bioCompany"],
                        "products": it["products"],
                    }
                    cells_out[it["code"]] = cell
                    row_cells.append(it["code"])
                    x += w
                rows_out.append({"row": row_idx, "y": y, "height": row_h, "type": row_type, "cells": row_cells})
                y += row_h
                row_idx += 1

        cap_height = None
        if block_height:
            leftover = float(block_height) - y
            if leftover > 1:
                cap_height = leftover
                rows_out.append({
                    "row": row_idx, "y": y, "height": leftover, "type": "cap",
                    "cells": [], "label": "Hộp gỗ ốp trên (trang trí)",
                })

        blocks_out[code] = {
            "code": code,
            "standalone": False,
            "name": (khoi or {}).get("name"),
            "mountPosition": (khoi or {}).get("mountPosition"),
            "width": block_width,
            "height": block_height or y,
            "depth": block_depth,
            "rows": rows_out,
        }

    return blocks_out, cells_out


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    khoi_blocks = read_khoi_sheet(wb)
    chia_o_entries = read_chia_o_sheet(wb)
    blocks, cells = build_blocks(khoi_blocks, chia_o_entries)

    out = {
        "generatedFrom": "Table.xlsx",
        "khoiBlocksRaw": khoi_blocks,
        "blocks": blocks,
        "cells": cells,
    }

    JSON_PATH.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")

    total_cells = sum(1 for c in cells.values() if c["kind"] != "cap")
    mapped = sum(1 for c in cells.values() if c.get("bizCode"))
    sys.stdout.write(
        "OK: %d blocks, %d cells (%d da anh xa doanh nhan) -> %s\n"
        % (len(blocks), total_cells, mapped, JSON_PATH.name)
    )


if __name__ == "__main__":
    main()
